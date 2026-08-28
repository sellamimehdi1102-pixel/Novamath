"""
Génération des exports Analytics (CSV/Excel/PDF) et des rapports enregistrés
(/admin/analytics — sections Export/Rapports). Réutilise EXCLUSIVEMENT les
agrégats déjà calculés par admin_analytics_service.py (overview/charts/
rankings) — jamais un recalcul séparé, jamais un accès direct à db.py pour
une statistique (Partie 13 : "ne jamais recalculer plusieurs fois les mêmes
statistiques").

Architecture stricte :

    server.py  →  admin_analytics_export_service.py  →  admin_analytics_service.py

jamais l'inverse. Ce module NE modifie JAMAIS le chatbot/Stripe/les
abonnements/les quotas/le routage IA — il ne fait que mettre en forme, dans
trois formats de fichier, des données déjà réelles et déjà validées ailleurs.

── Rapports enregistrés vs export ponctuel ─────────────────────────────────
Un export ponctuel (bouton "Exporter" de la page Analytics) est généré et
streamé directement, jamais écrit sur disque. Un "Rapport" (section Rapports,
quotidien/hebdomadaire/mensuel/annuel/personnalisé) est en plus ÉCRIT sur
disque (REPORTS_DIR) pour rester téléchargeable plus tard — l'historique
complet (date/heure/admin/type/format/taille/durée) est tenu par
admin_ai_log_service.py (table déjà générique, réutilisée telle quelle,
jamais une seconde table de log créée ici, voir sa docstring)."""
import csv
import io
import uuid
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.shapes import Drawing
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

import admin_ai_log_service
import admin_analytics_service as analytics
import db

REPORTS_DIR = db.DATA_DIR / "analytics_reports"

_BRAND_HEX = "7C5CFF"  # même violet que --accent (voir webapp/static/css/tokens.css)
_HEADER_HEX = "4B5563"

REPORT_TYPES = {
    "quotidien": "aujourd_hui", "hebdomadaire": "7j", "mensuel": "30j",
    "annuel": "cette_annee", "personnalise": "personnalise",
}
REPORT_TYPE_LABELS = {
    "quotidien": "Rapport quotidien", "hebdomadaire": "Rapport hebdomadaire",
    "mensuel": "Rapport mensuel", "annuel": "Rapport annuel", "personnalise": "Rapport personnalisé",
}

_CATEGORY_OVERVIEW_KEY = {
    "Utilisateurs": "users", "Chatbot": "chatbot", "IA": "ai",
    "Support": "support", "Abonnements": "subscriptions",
}
_CATEGORY_FIELDS = {
    "Utilisateurs": [
        ("Utilisateurs inscrits", "total_registered"), ("Nouveaux utilisateurs", "new_users"),
        ("Utilisateurs actifs", "active_users"), ("Connexions", "logins"),
    ],
    "Chatbot": [
        ("Conversations", "conversations"), ("Messages utilisateur", "user_messages"),
        ("Réponses assistant", "assistant_messages"),
    ],
    "IA": [
        ("Appels IA", "calls"), ("Tokens", "tokens"), ("Coût ($)", "cost_usd"),
        ("Latence moyenne (ms)", "avg_latency_ms"), ("Taux de succès (%)", "success_rate_pct"),
        ("Erreurs", "errors"), ("Fallbacks", "fallbacks"),
    ],
    "Support": [("Tickets créés", "tickets_created"), ("Ouverts", "open"), ("Fermés", "closed")],
    "Abonnements": [("Free", "free"), ("Premium", "premium"), ("Ultra", "ultra")],
}


def _timestamp():
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def _metric_display(metric):
    """(valeur affichable, remarque) à partir du contrat {available,value,
    reason/note} — jamais un 0/"" fabriqué : "Non disponible" + la raison
    exacte quand available=False."""
    if not metric or not metric.get("available"):
        reason = (metric or {}).get("reason") or ""
        return "Non disponible", reason
    value = metric["value"]
    if isinstance(value, (dict, list)):
        value = str(value)
    return value, metric.get("note", "")


def _category_rows(overview, category):
    section = overview[_CATEGORY_OVERVIEW_KEY[category]]
    return [(label, section[key]) for label, key in _CATEGORY_FIELDS[category]]


def collect(window="30j", date_from=None, date_to=None):
    """Point d'entrée UNIQUE : une seule évaluation de overview()/charts()/
    rankings(), réutilisée par les trois formats — jamais recalculée deux
    fois pour un même export."""
    overview, error = analytics.overview(window, date_from, date_to)
    if error:
        return None, error
    charts, _ = analytics.charts(window, date_from, date_to)
    rankings, _ = analytics.rankings(window, date_from, date_to, limit=10)
    return {"overview": overview, "charts": charts, "rankings": rankings}, None


# ── CSV ───────────────────────────────────────────────────────────────────────
def build_csv(window="30j", date_from=None, date_to=None):
    data, error = collect(window, date_from, date_to)
    if error:
        return None, None, error
    overview = data["overview"]
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["NovaMath — Export Analytics"])
    writer.writerow(["Période", overview["window"]["label"], overview["window"]["since"], overview["window"]["until"] or ""])
    writer.writerow([])
    for category in _CATEGORY_FIELDS:
        writer.writerow([category])
        writer.writerow(["Métrique", "Valeur", "Remarque"])
        for label, metric in _category_rows(overview, category):
            value, note = _metric_display(metric)
            writer.writerow([label, value, note])
        writer.writerow([])
    content = buf.getvalue().encode("utf-8-sig")
    filename = f"novamath_analytics_{window}_{_timestamp()}.csv"
    return content, filename, None


# ── Excel (openpyxl) ───────────────────────────────────────────────────────────
def _write_kpi_sheet(ws, title, rows):
    ws.merge_cells("A1:C1")
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=_BRAND_HEX)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    header_row = 3
    for col, header in enumerate(("Métrique", "Valeur", "Remarque"), start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_HEADER_HEX)
        cell.alignment = Alignment(horizontal="left")

    row = header_row + 1
    for label, metric in rows:
        value, note = _metric_display(metric)
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=3, value=note)
        row += 1

    if row > header_row + 1:
        ws.auto_filter.ref = f"A{header_row}:C{row - 1}"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 64


def build_excel(window="30j", date_from=None, date_to=None):
    data, error = collect(window, date_from, date_to)
    if error:
        return None, None, error
    overview = data["overview"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vue générale"
    general_rows = [
        (f"{category} — {label}", metric)
        for category in _CATEGORY_FIELDS
        for label, metric in _category_rows(overview, category)
    ]
    _write_kpi_sheet(ws, f"NovaMath Analytics — {overview['window']['label']}", general_rows)

    for category in _CATEGORY_FIELDS:
        sheet = wb.create_sheet(category)
        _write_kpi_sheet(sheet, category, _category_rows(overview, category))

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"novamath_analytics_{window}_{_timestamp()}.xlsx"
    return buf.getvalue(), filename, None


# ── PDF (reportlab) ────────────────────────────────────────────────────────────
def _series_drawing(chart_cls, series, color_hex, width=440, height=160):
    """Graphique vectoriel simple (courbe ou barres) à partir d'une série
    {"date","value"} déjà calculée par admin_analytics_service.charts() —
    jamais une capture d'écran du graphique Chart.js (rendu côté navigateur,
    inaccessible depuis le serveur) : ici un second rendu, RÉEL, des mêmes
    données, avec une bibliothèque adaptée à un document PDF."""
    points = [p for p in (series.get("value") if series and series.get("available") else []) if p["value"] is not None]
    if not points:
        return None
    drawing = Drawing(width, height)
    chart = chart_cls()
    chart.x, chart.y = 40, 20
    chart.width, chart.height = width - 60, height - 40
    values = [p["value"] for p in points]
    chart.data = [values]
    chart.categoryAxis.categoryNames = [p["date"][5:] for p in points]  # "MM-JJ"
    chart.categoryAxis.labels.angle = 45
    chart.categoryAxis.labels.fontSize = 6
    chart.valueAxis.valueMin = 0
    if hasattr(chart, "lines"):
        chart.lines[0].strokeColor = colors.HexColor(f"#{color_hex}")
        chart.lines[0].strokeWidth = 2
    if hasattr(chart, "bars"):
        chart.bars[0].fillColor = colors.HexColor(f"#{color_hex}")
    drawing.add(chart)
    return drawing


def _kpi_table(rows, styles):
    data = [["Métrique", "Valeur", "Remarque"]]
    for label, metric in rows:
        value, note = _metric_display(metric)
        data.append([Paragraph(label, styles["small"]), Paragraph(str(value), styles["small"]), Paragraph(note, styles["small"])])
    table = Table(data, colWidths=[5.5 * cm, 3.5 * cm, 7 * cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{_HEADER_HEX}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
    ]))
    return table


def _auto_summary(overview):
    """Résumé automatique — une phrase construite EXCLUSIVEMENT à partir de
    valeurs réelles déjà calculées (jamais un texte fabriqué) ; une métrique
    non disponible est simplement omise de la phrase, jamais remplacée par
    une estimation."""
    parts = []
    users = overview["users"]["new_users"]
    if users["available"]:
        parts.append(f"{users['value']} nouveau(x) utilisateur(s)")
    conv = overview["chatbot"]["conversations"]
    if conv["available"]:
        parts.append(f"{conv['value']} conversation(s)")
    calls = overview["ai"]["calls"]
    if calls["available"]:
        success = overview["ai"]["success_rate_pct"]
        success_txt = f" (taux de succès {success['value']}%)" if success["available"] else ""
        parts.append(f"{calls['value']} appel(s) IA{success_txt}")
    tickets = overview["support"]["tickets_created"]
    if tickets["available"]:
        parts.append(f"{tickets['value']} ticket(s) de support créé(s)")
    if not parts:
        return "Aucune activité mesurable n'a été enregistrée sur cette période."
    return "Sur la période sélectionnée, NovaMath a enregistré " + ", ".join(parts) + "."


def _page_footer(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#9CA3AF"))
    canvas.drawRightString(A4[0] - 2 * cm, 1.2 * cm, f"Page {doc.page}")
    canvas.drawString(2 * cm, 1.2 * cm, "NovaMath — Rapport Analytics")
    canvas.restoreState()


def build_pdf(window="30j", date_from=None, date_to=None, report_type_label=None):
    data, error = collect(window, date_from, date_to)
    if error:
        return None, None, error
    overview, charts_data, rankings = data["overview"], data["charts"], data["rankings"]

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="small", fontSize=8, leading=10))
    brand_style = ParagraphStyle(name="brand", fontSize=22, textColor=colors.HexColor(f"#{_BRAND_HEX}"), fontName="Helvetica-Bold")
    title_style = ParagraphStyle(name="title", fontSize=14, spaceAfter=4, fontName="Helvetica-Bold")
    meta_style = ParagraphStyle(name="meta", fontSize=9, textColor=colors.HexColor("#6B7280"))
    section_style = ParagraphStyle(name="section", fontSize=11, spaceBefore=14, spaceAfter=6, fontName="Helvetica-Bold")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm, topMargin=2 * cm, bottomMargin=2 * cm,
    )
    story = [
        Paragraph("NovaMath", brand_style),
        Spacer(1, 4),
        Paragraph(report_type_label or "Rapport Analytics", title_style),
        Paragraph(f"Généré le {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')} UTC", meta_style),
        Paragraph(f"Période : {overview['window']['label']} ({overview['window']['since']} → {overview['window']['until'] or 'maintenant'})", meta_style),
        Spacer(1, 10),
        Paragraph("Résumé", section_style),
        Paragraph(_auto_summary(overview), styles["Normal"]),
    ]

    for category in _CATEGORY_FIELDS:
        story.append(Paragraph(category, section_style))
        story.append(_kpi_table(_category_rows(overview, category), styles))

    users_daily = _series_drawing(HorizontalLineChart, charts_data["users"]["daily"], "3B82F6")
    if users_daily:
        story.append(Paragraph("Croissance quotidienne des utilisateurs", section_style))
        story.append(users_daily)

    ai_tokens = _series_drawing(VerticalBarChart, charts_data["ai"]["tokens_per_day"], "22C55E")
    if ai_tokens:
        story.append(Paragraph("Consommation de tokens IA par jour", section_style))
        story.append(ai_tokens)

    if rankings["ai"]["top_providers"]["available"]:
        story.append(Paragraph("Top fournisseurs IA", section_style))
        rows = [["Fournisseur", "Requêtes", "Tokens", "Coût ($)"]]
        for r in rankings["ai"]["top_providers"]["value"][:10]:
            rows.append([r["provider_key"], str(r["requests"]), str(r["total_tokens"]), f"{r['estimated_cost']:.4f}"])
        table = Table(rows, colWidths=[5 * cm, 3 * cm, 3 * cm, 3 * cm])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{_HEADER_HEX}")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ]))
        story.append(table)

    doc.build(story, onFirstPage=_page_footer, onLaterPages=_page_footer)
    filename = f"novamath_analytics_{window}_{_timestamp()}.pdf"
    return buf.getvalue(), filename, None


# ── Rapports enregistrés (écriture sur disque + métadonnées) ────────────────
def generate_report(report_type, export_format, date_from=None, date_to=None):
    """Génère un rapport (CSV/Excel/PDF) pour `report_type`
    (quotidien/hebdomadaire/mensuel/annuel/personnalisé), l'écrit dans
    REPORTS_DIR et renvoie (filename, content_bytes, duration_ms, error).
    Ne journalise rien ici : c'est à l'appelant (server.py) d'enregistrer
    l'action dans le journal administrateur (Partie 12), avec l'acteur réel."""
    import time as _time

    if report_type not in REPORT_TYPES:
        return None, None, None, f"Type de rapport invalide : {report_type!r}."
    window = REPORT_TYPES[report_type]

    started = _time.perf_counter()
    builders = {"csv": build_csv, "xlsx": build_excel, "pdf": build_pdf}
    builder = builders.get(export_format)
    if builder is None:
        return None, None, None, f"Format invalide : {export_format!r} (csv/xlsx/pdf attendus)."

    if export_format == "pdf":
        content, _filename, error = builder(window, date_from, date_to, REPORT_TYPE_LABELS[report_type])
    else:
        content, _filename, error = builder(window, date_from, date_to)
    if error:
        return None, None, None, error

    duration_ms = round((_time.perf_counter() - started) * 1000)
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"{report_type}_{uuid.uuid4().hex[:8]}_{_timestamp()}.{export_format}"
    (REPORTS_DIR / filename).write_bytes(content)
    return filename, len(content), duration_ms, None


def resolve_report_path(filename):
    """None si `filename` ne correspond à aucun fichier réel dans
    REPORTS_DIR — jamais un chemin arbitraire accepté tel quel (même
    politique que backup_service.resolve_backup_path)."""
    path = REPORTS_DIR / filename
    if "/" in filename or "\\" in filename or not path.is_file():
        return None
    try:
        path.relative_to(REPORTS_DIR)
    except ValueError:
        return None
    return path


# ── Journalisation (Partie 12 : chaque export DOIT être enregistré) ─────────
# Même convention "*_logged" que admin_ai_service.py/settings_service.py :
# la fonction d'origine reste pure/testable isolément, la version "_logged"
# ajoute UNIQUEMENT l'écriture dans le journal administrateur (ai_admin_log,
# déjà générique — jamais une seconde table de log créée ici).
_EXPORT_BUILDERS = {"csv": (build_csv, "analytics_export_csv"), "xlsx": (build_excel, "analytics_export_excel"),
                    "pdf": (build_pdf, "analytics_export_pdf")}


def export_logged(actor, export_format, window="30j", date_from=None, date_to=None):
    """Renvoie (content_bytes, filename, error) — journalise TOUJOURS une
    entrée (résultat "success" ou "error"), jamais un export silencieux."""
    import time as _time

    entry = _EXPORT_BUILDERS.get(export_format)
    if entry is None:
        return None, None, f"Format invalide : {export_format!r} (csv/xlsx/pdf attendus)."
    builder, action = entry

    started = _time.perf_counter()
    content, filename, error = builder(window, date_from, date_to)
    duration_ms = round((_time.perf_counter() - started) * 1000)

    if error:
        admin_ai_log_service.record(actor, action=action, provider_name=window, result="error", error_message=error)
        return None, None, error

    admin_ai_log_service.record(
        actor, action=action, provider_name=window, result="success",
        new_values={"filename": filename, "format": export_format, "size_bytes": len(content), "duration_ms": duration_ms},
    )
    return content, filename, None


def generate_report_logged(actor, report_type, export_format, date_from=None, date_to=None):
    """Renvoie (filename, error) — même politique de journalisation
    systématique que export_logged() ci-dessus, action "analytics_report"."""
    filename, size_bytes, duration_ms, error = generate_report(report_type, export_format, date_from, date_to)
    if error:
        admin_ai_log_service.record(
            actor, action="analytics_report", provider_name=report_type, result="error", error_message=error,
        )
        return None, error
    admin_ai_log_service.record(
        actor, action="analytics_report", provider_name=report_type, result="success",
        new_values={
            "filename": filename, "format": export_format, "report_type": report_type,
            "size_bytes": size_bytes, "duration_ms": duration_ms,
        },
    )
    return filename, None


def list_export_history(limit=50):
    """Historique fusionné des 4 actions Analytics (exports + rapports),
    triable par date — réutilise admin_ai_log_service.list_log() tel quel,
    jamais une seconde requête/table d'historique créée ici (Partie 6)."""
    actions = ("analytics_export_csv", "analytics_export_excel", "analytics_export_pdf", "analytics_report")
    entries = []
    for action in actions:
        entries.extend(admin_ai_log_service.list_log(action=action, page_size=limit)["items"])
    entries.sort(key=lambda e: e["created_at"], reverse=True)
    return entries[:limit]
